"""Export functionality for subwoofer simulation data."""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import logging
import json
from datetime import datetime

logger = logging.getLogger(__name__)


class ProjectExporter:
    """Handles exporting simulation data to various formats."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def export_to_excel(
        self,
        file_path: Union[str, Path],
        project_data: Dict[str, Any],
        include_metadata: bool = True
    ) -> None:
        """Export project data to Excel format.
        
        Args:
            file_path: Path to save Excel file
            project_data: Dictionary containing all project data
            include_metadata: Whether to include metadata sheet
        """
        file_path = Path(file_path)
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Export sources
                if 'sources' in project_data:
                    sources_df = self._sources_to_dataframe(project_data['sources'])
                    sources_df.to_excel(writer, sheet_name='Sources', index=False)
                    self._format_sources_sheet(writer, 'Sources')
                
                # Export room geometry
                if 'room_vertices' in project_data:
                    room_df = pd.DataFrame(project_data['room_vertices'], columns=['X', 'Y'])
                    room_df.to_excel(writer, sheet_name='Room', index=False)
                
                # Export target areas
                if 'target_areas' in project_data:
                    target_df = self._areas_to_dataframe(project_data['target_areas'], 'target')
                    target_df.to_excel(writer, sheet_name='Target_Areas', index=False)
                
                # Export avoidance areas
                if 'avoidance_areas' in project_data:
                    avoidance_df = self._areas_to_dataframe(project_data['avoidance_areas'], 'avoidance')
                    avoidance_df.to_excel(writer, sheet_name='Avoidance_Areas', index=False)
                
                # Export sub placement areas
                if 'sub_placement_areas' in project_data:
                    placement_df = self._areas_to_dataframe(project_data['sub_placement_areas'], 'placement')
                    placement_df.to_excel(writer, sheet_name='Placement_Areas', index=False)
                
                # Export array groups
                if 'array_groups' in project_data:
                    groups_df = self._groups_to_dataframe(project_data['array_groups'])
                    groups_df.to_excel(writer, sheet_name='Array_Groups', index=False)
                
                # Export simulation parameters
                if 'simulation_params' in project_data:
                    params_df = self._params_to_dataframe(project_data['simulation_params'])
                    params_df.to_excel(writer, sheet_name='Parameters', index=False)
                
                # Export optimization results
                if 'optimization_results' in project_data:
                    optim_df = self._optimization_to_dataframe(project_data['optimization_results'])
                    optim_df.to_excel(writer, sheet_name='Optimization', index=False)
                
                # Export metadata
                if include_metadata:
                    metadata_df = self._create_metadata_dataframe(project_data)
                    metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
            
            self.logger.info(f"Project exported to {file_path}")
            
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _sources_to_dataframe(self, sources: np.ndarray) -> pd.DataFrame:
        """Convert sources array to DataFrame."""
        data = []
        for i, source in enumerate(sources):
            data.append({
                'ID': i + 1,
                'X_Position': source['x'],
                'Y_Position': source['y'],
                'SPL_at_1m_dB': 20 * np.log10(source['pressure_val_at_1m_relative_to_pref']),
                'Gain_dB': 20 * np.log10(source['gain_lin']),
                'Angle_degrees': np.degrees(source['angle']),
                'Delay_ms': source['delay_ms'],
                'Polarity': source['polarity']
            })
        
        return pd.DataFrame(data)
    
    def _areas_to_dataframe(self, areas: List[Dict], area_type: str) -> pd.DataFrame:
        """Convert areas list to DataFrame."""
        data = []
        for area in areas:
            vertices = area.get('vertices', [])
            for i, vertex in enumerate(vertices):
                data.append({
                    'Area_ID': area.get('id', 0),
                    'Area_Type': area_type,
                    'Vertex_Index': i,
                    'X': vertex[0],
                    'Y': vertex[1],
                    'Min_SPL': area.get('min_spl', ''),
                    'Max_SPL': area.get('max_spl', ''),
                    'Name': area.get('name', '')
                })
        
        return pd.DataFrame(data)
    
    def _groups_to_dataframe(self, groups: Dict[str, Any]) -> pd.DataFrame:
        """Convert array groups to DataFrame."""
        data = []
        for group_id, group_data in groups.items():
            data.append({
                'Group_ID': group_id,
                'Type': group_data.get('tipo', ''),
                'Frequency': group_data.get('freq', ''),
                'Radius': group_data.get('radius', ''),
                'Mode': group_data.get('mode', ''),
                'Steering_Angle': group_data.get('steering_angle', ''),
                'Coverage_Angle': group_data.get('coverage_angle', ''),
                'Start_Angle': group_data.get('start_angle', ''),
                'Source_IDs': ','.join(map(str, group_data.get('source_ids', [])))
            })
        
        return pd.DataFrame(data)
    
    def _params_to_dataframe(self, params: Dict[str, Any]) -> pd.DataFrame:
        """Convert simulation parameters to DataFrame."""
        data = []
        for key, value in params.items():
            data.append({
                'Parameter': key,
                'Value': str(value),
                'Type': type(value).__name__
            })
        
        return pd.DataFrame(data)
    
    def _optimization_to_dataframe(self, results: Dict[str, Any]) -> pd.DataFrame:
        """Convert optimization results to DataFrame."""
        data = []
        for key, value in results.items():
            data.append({
                'Metric': key,
                'Value': str(value),
                'Unit': self._get_unit_for_metric(key)
            })
        
        return pd.DataFrame(data)
    
    def _get_unit_for_metric(self, metric: str) -> str:
        """Get unit for optimization metric."""
        unit_map = {
            'fitness': 'score',
            'generation': 'count',
            'frequency': 'Hz',
            'spl': 'dB',
            'delay': 'ms',
            'gain': 'dB',
            'angle': 'degrees',
            'position': 'm'
        }
        
        for key, unit in unit_map.items():
            if key in metric.lower():
                return unit
        
        return ''
    
    def _create_metadata_dataframe(self, project_data: Dict[str, Any]) -> pd.DataFrame:
        """Create metadata DataFrame."""
        metadata = {
            'Export_Date': datetime.now().isoformat(),
            'Project_Name': project_data.get('name', 'Unnamed Project'),
            'Description': project_data.get('description', ''),
            'Version': project_data.get('version', '1.0'),
            'Author': project_data.get('author', ''),
            'Software': 'Subwoofer Simulation Tool',
            'Number_of_Sources': len(project_data.get('sources', [])),
            'Number_of_Target_Areas': len(project_data.get('target_areas', [])),
            'Number_of_Avoidance_Areas': len(project_data.get('avoidance_areas', [])),
            'Number_of_Placement_Areas': len(project_data.get('sub_placement_areas', [])),
            'Number_of_Array_Groups': len(project_data.get('array_groups', {}))
        }
        
        data = [{'Property': key, 'Value': str(value)} for key, value in metadata.items()]
        return pd.DataFrame(data)
    
    def _format_sources_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Format the sources sheet with appropriate styling."""
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Set column widths
            column_widths = {
                'A': 5,   # ID
                'B': 12,  # X_Position
                'C': 12,  # Y_Position
                'D': 15,  # SPL_at_1m_dB
                'E': 10,  # Gain_dB
                'F': 15,  # Angle_degrees
                'G': 12,  # Delay_ms
                'H': 10   # Polarity
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Format headers
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            center_alignment = Alignment(horizontal="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Format numeric columns
            for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=8):
                for cell in row:
                    cell.alignment = center_alignment
                    if cell.column_letter in ['B', 'C', 'D', 'E', 'G']:
                        cell.number_format = '0.00'
            
        except Exception as e:
            self.logger.warning(f"Could not format Excel sheet: {e}")
    
    def export_to_json(
        self,
        file_path: Union[str, Path],
        project_data: Dict[str, Any],
        indent: int = 2
    ) -> None:
        """Export project data to JSON format.
        
        Args:
            file_path: Path to save JSON file
            project_data: Dictionary containing all project data
            indent: JSON indentation level
        """
        file_path = Path(file_path)
        
        try:
            # Convert numpy arrays to lists for JSON serialization
            serializable_data = self._make_json_serializable(project_data)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(serializable_data, f, indent=indent, ensure_ascii=False)
            
            self.logger.info(f"Project exported to JSON: {file_path}")
            
        except Exception as e:
            self.logger.error(f"Error exporting to JSON: {e}")
            raise
    
    def _make_json_serializable(self, obj: Any) -> Any:
        """Make object JSON serializable."""
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, dict):
            return {key: self._make_json_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._make_json_serializable(item) for item in obj]
        else:
            return obj
    
    def export_to_csv(
        self,
        file_path: Union[str, Path],
        data: Union[pd.DataFrame, np.ndarray, List],
        headers: Optional[List[str]] = None
    ) -> None:
        """Export data to CSV format.
        
        Args:
            file_path: Path to save CSV file
            data: Data to export
            headers: Optional column headers
        """
        file_path = Path(file_path)
        
        try:
            if isinstance(data, pd.DataFrame):
                data.to_csv(file_path, index=False, encoding='utf-8')
            elif isinstance(data, np.ndarray):
                df = pd.DataFrame(data, columns=headers)
                df.to_csv(file_path, index=False, encoding='utf-8')
            elif isinstance(data, list):
                df = pd.DataFrame(data, columns=headers)
                df.to_csv(file_path, index=False, encoding='utf-8')
            else:
                raise ValueError(f"Unsupported data type: {type(data)}")
            
            self.logger.info(f"Data exported to CSV: {file_path}")
            
        except Exception as e:
            self.logger.error(f"Error exporting to CSV: {e}")
            raise
    
    def export_spl_map(
        self,
        file_path: Union[str, Path],
        X_grid: np.ndarray,
        Y_grid: np.ndarray,
        SPL_grid: np.ndarray,
        frequency: float
    ) -> None:
        """Export SPL map data to CSV.
        
        Args:
            file_path: Path to save CSV file
            X_grid: X coordinate grid
            Y_grid: Y coordinate grid
            SPL_grid: SPL values grid
            frequency: Frequency of the SPL map
        """
        file_path = Path(file_path)
        
        try:
            # Flatten grids
            x_flat = X_grid.flatten()
            y_flat = Y_grid.flatten()
            spl_flat = SPL_grid.flatten()
            
            # Create DataFrame
            df = pd.DataFrame({
                'X': x_flat,
                'Y': y_flat,
                'SPL_dB': spl_flat,
                'Frequency_Hz': frequency
            })
            
            # Remove invalid SPL values
            df = df[df['SPL_dB'] > -200]
            
            df.to_csv(file_path, index=False, encoding='utf-8')
            
            self.logger.info(f"SPL map exported to CSV: {file_path}")
            
        except Exception as e:
            self.logger.error(f"Error exporting SPL map: {e}")
            raise
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats."""
        return ['xlsx', 'json', 'csv']
    
    def validate_export_data(self, project_data: Dict[str, Any]) -> bool:
        """Validate project data before export.
        
        Args:
            project_data: Dictionary containing project data
            
        Returns:
            True if valid, False otherwise
        """
        try:
            # Check for required fields
            if 'sources' not in project_data:
                self.logger.error("No sources data found")
                return False
            
            # Validate sources
            sources = project_data['sources']
            if not isinstance(sources, np.ndarray) or len(sources) == 0:
                self.logger.error("Invalid sources data")
                return False
            
            # Check for required source fields
            required_fields = ['x', 'y', 'pressure_val_at_1m_relative_to_pref', 
                             'gain_lin', 'angle', 'delay_ms', 'polarity']
            
            for field in required_fields:
                if field not in sources.dtype.names:
                    self.logger.error(f"Missing required source field: {field}")
                    return False
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error validating export data: {e}")
            return False